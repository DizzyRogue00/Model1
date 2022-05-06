import gurobipy
import pandas as pd


def assignment(cost_matrix):
    # 保存行列标签
    index = cost_matrix.index
    columns = cost_matrix.columns

    # 创建模型
    model = gurobipy.Model('Assignment')
    x = model.addVars(index, columns, vtype=gurobipy.GRB.BINARY, name='zc')
    model.update()

    # 设置目标函数
    model.setObjective(gurobipy.quicksum(x[i, j] * cost_matrix.at[i, j] for i in index for j in columns))

    # 添加约束条件
    model.addConstr(gurobipy.quicksum(x[i, j] for i in index for j in columns) == min([len(index), len(columns)]))
    model.addConstrs(gurobipy.quicksum(x[i, j] for j in columns) <= 1 for i in index)
    model.addConstrs(gurobipy.quicksum(x[i, j] for i in index) <= 1 for j in columns)

    # 执行最优化
    model.optimize()

    # 输出信息
    result = cost_matrix * 0
    if model.status == gurobipy.GRB.Status.OPTIMAL:
        solution = [k for k, v in model.getAttr('x', x).items() if v == 1]
    print(model.getAttr('X',x))
    print(model.getVars())
    for i, j in solution:
        print(f"{i} -> {j}：{cost_matrix.at[i, j]}")
        result.at[i, j] = 1
    return result


if __name__ == '__main__':
    cost_matrix = pd.DataFrame(
        [[4, 8, 7, 15, 12], [7, 9, 17, 14, 10], [6, 9, 12, 8, 7], [6, 7, 14, 6, 10], [6, 9, 12, 10, 6],
         [5, 8, 13, 11, 10]],
        index=['A1', 'A2', 'A3', 'A4', 'A5', 'A6'], columns=['B1', 'B2', 'B3', 'B4', 'B5'])

    assignment(cost_matrix)
import pathlib
from openpyxl import load_workbook
'''
with pd.ExcelWriter('test.xlsx', engine='openpyxl') as writer:
    d.to_excel(writer,sheet_name='sheet1')
with pd.ExcelWriter('test.xlsx', engine='openpyxl',mode='a') as writer:
    df2.to_excel(writer,sheet_name='sheet2')
path = pathlib.Path('test.xlsx')
if path.exists():
    with pd.ExcelWriter('test.xlsx', engine='openpyxl',if_sheet_exists='') as writer:
    writer = pd.ExcelWriter('test.xlsx', engine='openpyxl')
    book = load_workbook(writer.path)
    writer.book = book
    d.to_excel(excel_writer=writer, sheet_name="sheet1")
    writer.save()
    writer.close()
else:
    writer = pd.ExcelWriter('test.xlsx', engine='openpyxl')
    d.to_excel(writer, sheet_name="sheet1")
    writer.save()
    writer.close()

if path.exists():
    writer = pd.ExcelWriter('test.xlsx')
    book = load_workbook(writer.path)
    writer.book = book
    df2.to_excel(excel_writer=writer, sheet_name="sheet2")
    writer.save()
    writer.close()
else:
    writer = pd.ExcelWriter('test.xlsx')
    df2.to_excel(writer, sheet_name="sheet2")
    writer.save()
    writer.close()
'''

'''
import traceback
import logging
 
logging.basicConfig(filename='log.txt', level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
with open('log.txt',"a") as file:
    try:
        file.write('try...\n')
        r = 10 / 0
        file.write('result:{}'.format(r))
    except ZeroDivisionError as e:
        file.write('except: {}\n'.format(e))
        file.write(traceback.format_exc())
    finally:
        file.write('finally...\n')
        file.write('END\n')    
'''

import time
import datetime
t1=time.perf_counter()
t2=time.perf_counter_ns()
t3=time.clock()
t4=time.process_time()
t5=time.process_time_ns()
start_time=datetime.datetime.now()
start=time.time()

zc_sum=0
for i in range(100*100):
    zc_sum+=i
t11=time.perf_counter()
t22=time.perf_counter_ns()
t33=time.clock()
t44=time.process_time()
t55=time.process_time_ns()
end_time=datetime.datetime.now()
end=time.time()

print('程序运行时间:%s' % (t11 - t1))#s
print('程序运行时间:%s' % ((t22 - t2)/1000/1000))#ms
print('程序运行时间:%s' % (t33 - t3))#s cpu time
print('程序运行时间:%s' % (t44 - t4))#s cpu time
print('程序运行时间:%s' % (t55 - t5))#s cpu time
print('程序运行时间:%s' % (end_time-start_time).microseconds)#micro seconds
print('程序运行时间:%s' % ((end-start)*1000))#ms
